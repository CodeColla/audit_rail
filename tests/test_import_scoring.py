"""M6 — xlsx import wizard, mapping proposals, scoring, and export."""

import io
import json
import uuid

from openpyxl import Workbook, load_workbook
from sqlalchemy import text as sqltext

from tests.conftest import token


def _make_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.append(["S.No", "Domain", "Control Question"])
    ws.append(["1", "Access Management", "Do you enforce strong password policies?"])
    ws.append(["2", "Access Management", "Do you support MFA and SSO authentication?"])
    ws.append(["3", "Governance", "Do you have an information security policy?"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _seed_controls(engine, tenant_id):
    with engine.begin() as c:
        dom = c.execute(sqltext("SELECT id FROM domains WHERE code='AM' AND tenant_id=:t"),
                        {"t": tenant_id}).scalar()
        for code, stmt in [("IMP 1.a", "Strong password policy"),
                           ("IMP 2.a", "Strong authentication SSO MFA")]:
            c.execute(sqltext(
                "INSERT INTO controls (id,tenant_id,domain_id,code,statement,lifecycle,"
                "applicability,status,created_at,updated_at) VALUES "
                "(:i,:t,:d,:code,:s,'one_time','applicable','active',"
                "'2026-07-11T00:00:00Z','2026-07-11T00:00:00Z')"),
                {"i": str(uuid.uuid4()), "t": tenant_id, "d": dom, "code": code, "s": stmt})


def test_import_proposals_scoring_export(app_client):
    from api.core.database import engine
    with engine.connect() as c:
        tid = c.execute(sqltext("SELECT id FROM tenants WHERE slug='kiam'")).scalar()
    _seed_controls(engine, tid)
    tok = token(app_client, "member@kiam.example", "secret2")
    h = {"Authorization": f"Bearer {tok}"}

    # import a bank checklist — P7-S6b split this into preview (parse only) then commit
    # (write, given the rows the preview returned)
    xlsx_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    pv = app_client.post("/api/templates/import/preview", headers=h,
                        files={"file": ("hdfc.xlsx", _make_xlsx(), xlsx_type)})
    assert pv.status_code == 200, pv.text
    assert pv.json()["meta"]["number_column_detected"] is True
    rows = pv.json()["rows"]
    assert all(row["number"] for row in rows)

    r = app_client.post("/api/templates/import", headers=h,
        data={"bank_name": "HDFC", "rows": json.dumps(rows)},
        files={"file": ("hdfc.xlsx", _make_xlsx(), xlsx_type)})
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["questions"] == 3
    assert body["sections"] == 2      # Access Management + Governance
    assert body["proposals"] >= 2     # password + MFA questions mapped to AM controls
    tpl = body["template_id"]

    # proposals: password question mapped to the password control with confidence
    proposals = app_client.get(f"/api/templates/{tpl}/proposals", headers=h).json()
    pw = next(p for p in proposals if "password" in p["text"].lower())
    assert pw["code"] == "IMP 1.a"
    assert pw["status"] == "suggested" and pw["confidence"] > 0

    # bulk-confirm the high-confidence ones
    conf = app_client.post(f"/api/templates/{tpl}/proposals/confirm", headers=h,
                          json={"confirm_high_confidence": 0.3}).json()
    assert conf["confirmed"] >= 1
    again = app_client.get(f"/api/templates/{tpl}/proposals", headers=h).json()
    assert any(p["status"] == "confirmed" for p in again)

    # scoring config round-trips and drives the verdict
    cfg = app_client.get(f"/api/templates/{tpl}/scoring", headers=h).json()
    assert cfg["verdict_thresholds"][0]["verdict"] == "Satisfactory"

    # create an assessment, answer 1 yes + 1 no -> 50% -> Pending (default bands)
    aid = app_client.post("/api/assessments", headers=h,
                         json={"template_id": tpl, "title": "HDFC 2026"}).json()["id"]
    qs = app_client.get(f"/api/templates/{tpl}/questions", headers=h).json()
    app_client.put(f"/api/assessments/{aid}/responses/{qs[0]['id']}", headers=h,
                  json={"response_value": "yes"})
    app_client.put(f"/api/assessments/{aid}/responses/{qs[1]['id']}", headers=h,
                  json={"response_value": "no"})
    det = app_client.get(f"/api/assessments/{aid}", headers=h).json()
    assert det["score_pct"] == 50
    assert det["predicted_verdict"] == "Pending"

    # tighten thresholds so 50% now reads Conditional
    app_client.put(f"/api/templates/{tpl}/scoring", headers=h, json={"config": {
        "response_weights": {"yes": 1.0, "partial": 0.5, "no": 0.0, "na": None},
        "verdict_thresholds": [{"verdict": "Satisfactory", "min_pct": 80},
                               {"verdict": "Conditional", "min_pct": 40},
                               {"verdict": "Pending", "min_pct": 0}]}})
    det2 = app_client.get(f"/api/assessments/{aid}", headers=h).json()
    assert det2["predicted_verdict"] == "Conditional"

    # export answers as a real xlsx
    exp = app_client.get(f"/api/assessments/{aid}/export.xlsx", headers=h)
    assert exp.status_code == 200
    assert exp.content[:2] == b"PK"                      # xlsx = zip
    wb = load_workbook(io.BytesIO(exp.content))
    cells = [v for row in wb.active.iter_rows(values_only=True) for v in row]
    assert "YES" in cells                                # our answer made it in
    assert any(c and "password" in str(c).lower() for c in cells)


# ─────────────────────────────────────────────────────────── P7-S6b

def _make_xlsx_no_number_col():
    wb = Workbook()
    ws = wb.active
    ws.append(["Domain", "Control Question"])
    ws.append(["Access Management", "Do you enforce strong password policies?"])
    ws.append(["Governance", "Do you have an information security policy?"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_preview_writes_nothing_to_the_database(app_client):
    """The whole point of a preview: parse, don't commit. A second preview of the same file
    must not create a second template — because it must not create one at all."""
    from api.core.database import engine
    tok = token(app_client, "member@kiam.example", "secret2")
    h = {"Authorization": f"Bearer {tok}"}
    with engine.connect() as c:
        tid = c.execute(sqltext("SELECT id FROM tenants WHERE slug='kiam'")).scalar()
        before = c.execute(sqltext(
            "SELECT count(*) FROM templates WHERE tenant_id=:t"), {"t": tid}).scalar()

    xlsx_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    pv = app_client.post("/api/templates/import/preview", headers=h,
                        files={"file": ("hdfc.xlsx", _make_xlsx(), xlsx_type)})
    assert pv.status_code == 200, pv.text
    assert len(pv.json()["rows"]) == 3

    with engine.connect() as c:
        after = c.execute(sqltext(
            "SELECT count(*) FROM templates WHERE tenant_id=:t"), {"t": tid}).scalar()
    assert after == before


def test_preview_flags_a_missing_number_column(app_client):
    """The exact failure the issue named: a checklist with no Number column at all used to
    silently import every question unnumbered. The preview must say so explicitly, not
    leave the frontend to infer it from every row happening to be blank."""
    tok = token(app_client, "member@kiam.example", "secret2")
    h = {"Authorization": f"Bearer {tok}"}
    xlsx_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    pv = app_client.post("/api/templates/import/preview", headers=h,
                        files={"file": ("no-numbers.xlsx", _make_xlsx_no_number_col(), xlsx_type)})
    assert pv.status_code == 200, pv.text
    assert pv.json()["meta"]["number_column_detected"] is False
    assert all(not row["number"] for row in pv.json()["rows"])


def test_import_is_rejected_when_any_row_has_no_number(app_client):
    """The mandate itself: committing with a blank number anywhere must 400, whether that
    blank came from the file or was never fixed in the preview step. Defence in depth —
    this must hold even for a caller that skips the UI's own "fix it here" grid entirely."""
    tok = token(app_client, "member@kiam.example", "secret2")
    h = {"Authorization": f"Bearer {tok}"}
    xlsx_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    rows = [{"number": "1", "section": "Access", "text": "Q1?"},
           {"number": "", "section": "Access", "text": "Q2?"}]
    r = app_client.post("/api/templates/import", headers=h,
        data={"bank_name": "Blank Number Bank", "rows": json.dumps(rows)},
        files={"file": ("f.xlsx", _make_xlsx(), xlsx_type)})
    assert r.status_code == 400
    assert "2" in r.json()["detail"]                     # names the offending row


def test_import_commits_exactly_the_rows_it_was_given_not_a_reparse(app_client):
    """Committing must use the (possibly hand-edited) rows verbatim — not re-parse the file,
    which could silently disagree with what the user approved on screen."""
    tok = token(app_client, "member@kiam.example", "secret2")
    h = {"Authorization": f"Bearer {tok}"}
    xlsx_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # deliberately doesn't match _make_xlsx()'s real content — proves the FILE isn't re-parsed
    rows = [{"number": "9.z", "section": "Hand-edited", "text": "A hand-fixed question"}]
    r = app_client.post("/api/templates/import", headers=h,
        data={"bank_name": "Hand Edited Bank", "rows": json.dumps(rows)},
        files={"file": ("f.xlsx", _make_xlsx(), xlsx_type)})
    assert r.status_code == 201, r.text
    tpl = r.json()["template_id"]
    qs = app_client.get(f"/api/templates/{tpl}/questions", headers=h).json()
    assert len(qs) == 1
    assert qs[0]["number"] == "9.z"
    assert qs[0]["text"] == "A hand-fixed question"
