"""P4-S6 — Audits: re-mapping a question's control, evidence inherited from that control,
and the export's evidence column.

The two bugs this sprint fixed, both pinned below:
  • confirm_proposals accepted any control_id with no tenant check — a garbage or
    cross-tenant id hit the composite FK and raised an unhandled IntegrityError (500).
  • export_answers / questions_grid queried response_evidence with NO scoping at all,
    walking every row in the database on every call.
"""

import uuid

from sqlalchemy import text as sqltext

from tests.conftest import token
from tests.test_assessments import _seed


def _h(client, who="admin@kiam.example", pw="secret1"):
    return {"Authorization": f"Bearer {token(client, who, pw)}"}


def _tid(engine):
    with engine.connect() as c:
        return c.execute(sqltext("SELECT id FROM tenants WHERE slug='kiam'")).scalar()


def _assessment(app_client, h, tpl, title="Remap audit"):
    return app_client.post("/api/assessments", headers=h,
                           json={"template_id": tpl, "title": title}).json()["id"]


# ---------------------------------------------------------------- confirm_proposals guard

def test_confirm_proposals_rejects_unknown_control_id(app_client):
    """Was an unhandled 500 from the composite FK; must be a readable 400."""
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-cp1")
    r = app_client.post(f"/api/templates/{ids['tpl']}/proposals/confirm", headers=h, json={
        "decisions": [{"question_id": ids["q1"], "action": "confirm",
                       "control_id": str(uuid.uuid4())}]})
    assert r.status_code == 400
    assert "organisation" in r.json()["detail"]


def test_confirm_proposals_accepts_a_same_tenant_control_id(app_client):
    """The remap path itself works — this is what the Import review screen now sends."""
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-cp2")
    r = app_client.post(f"/api/templates/{ids['tpl']}/proposals/confirm", headers=h, json={
        "decisions": [{"question_id": ids["q1"], "action": "confirm",
                       "control_id": ids["ctl2"]}]})
    assert r.status_code == 200

    props = app_client.get(f"/api/templates/{ids['tpl']}/proposals", headers=h).json()
    q1 = next(p for p in props if p["question_id"] == ids["q1"])
    assert q1["control_id"] == ids["ctl2"] and q1["status"] == "confirmed"


# ---------------------------------------------------------------- PATCH .../mapping

def test_remap_points_a_question_at_a_different_control(app_client):
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-rm1")
    aid = _assessment(app_client, h, ids["tpl"])

    d = app_client.get(f"/api/assessments/{aid}/responses/{ids['q1']}", headers=h).json()
    assert d["mapped_control"]["code"] == ids["code1"]

    r = app_client.patch(f"/api/assessments/{aid}/responses/{ids['q1']}/mapping",
                         headers=h, json={"control_id": ids["ctl2"]})
    assert r.status_code == 200, r.text

    d2 = app_client.get(f"/api/assessments/{aid}/responses/{ids['q1']}", headers=h).json()
    assert d2["mapped_control"]["code"] == ids["code2"]


def test_remap_reports_but_does_not_rewrite_a_prefilled_answer(app_client):
    """Same "report, don't silently rewrite" rule as PATCHing a control's stock_response:
    response_revisions is an audit trail and a bank auditor may already have read the old
    answer, so the caller is told what is now stale instead of it changing under them."""
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-rm2")
    aid = _assessment(app_client, h, ids["tpl"])
    assert app_client.post(f"/api/assessments/{aid}/prefill", headers=h).json()["prefilled"] == 1

    before = app_client.get(f"/api/assessments/{aid}/responses/{ids['q1']}", headers=h).json()
    assert before["response"]["response_value"] == "yes"

    r = app_client.patch(f"/api/assessments/{aid}/responses/{ids['q1']}/mapping",
                         headers=h, json={"control_id": ids["ctl2"]})
    assert r.json() == {"ok": True, "was_prefilled": True}

    after = app_client.get(f"/api/assessments/{aid}/responses/{ids['q1']}", headers=h).json()
    assert after["response"]["response_value"] == "yes", "the saved answer must be untouched"


def test_remap_of_a_handwritten_answer_is_not_flagged_as_prefilled(app_client):
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-rm3")
    aid = _assessment(app_client, h, ids["tpl"])
    app_client.put(f"/api/assessments/{aid}/responses/{ids['q1']}", headers=h,
                   json={"response_value": "no", "comment": "written by a human"})
    r = app_client.patch(f"/api/assessments/{aid}/responses/{ids['q1']}/mapping",
                         headers=h, json={"control_id": ids["ctl2"]})
    assert r.json()["was_prefilled"] is False


def test_remap_rejects_unknown_control_and_foreign_question(app_client):
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-rm4")
    aid = _assessment(app_client, h, ids["tpl"])

    bad_ctl = app_client.patch(f"/api/assessments/{aid}/responses/{ids['q1']}/mapping",
                               headers=h, json={"control_id": str(uuid.uuid4())})
    assert bad_ctl.status_code == 400

    bad_q = app_client.patch(f"/api/assessments/{aid}/responses/{uuid.uuid4()}/mapping",
                             headers=h, json={"control_id": ids["ctl2"]})
    assert bad_q.status_code == 404


def test_remap_is_member_only(app_client):
    """An auditor guest can read the audit but must never rewrite the control library's
    relationship to it."""
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-rm5")
    aid = _assessment(app_client, h, ids["tpl"])
    inv = app_client.post(f"/api/assessments/{aid}/guests", headers=h, json={
        "email": f"auditor-{uuid.uuid4().hex[:6]}@pwc.example", "full_name": "A. Guest",
        "firm": "PwC"})
    gh = {"Authorization": f"Bearer {inv.json()['access_token']}"}
    r = app_client.patch(f"/api/assessments/{aid}/responses/{ids['q1']}/mapping",
                         headers=gh, json={"control_id": ids["ctl2"]})
    assert r.status_code == 403


def test_remap_requires_auth(app_client):
    assert app_client.patch(
        f"/api/assessments/{uuid.uuid4()}/responses/{uuid.uuid4()}/mapping",
        json={"control_id": "x"}).status_code == 401


# ---------------------------------------------------------------- inherited evidence

def test_control_evidence_is_inherited_by_its_mapped_questions(app_client):
    """P4-S5 let evidence attach to a CONTROL; P4-S6 surfaces it on every audit question
    mapped to that control — as a separate list, never merged into the question's own."""
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-inh")
    aid = _assessment(app_client, h, ids["tpl"])

    ev = app_client.post("/api/evidence", headers=h,
                         data={"title": "ISP PDF", "evidence_type": "POLICY_DOC"},
                         files={"file": ("isp.pdf", b"%PDF-1.4 x", "application/pdf")})
    assert ev.status_code == 201, ev.text
    eid = ev.json()["id"]
    assert app_client.post(f"/api/library/controls/{ids['ctl1']}/evidence", headers=h,
                           json={"evidence_id": eid}).status_code == 201

    d = app_client.get(f"/api/assessments/{aid}/responses/{ids['q1']}", headers=h).json()
    assert [e["id"] for e in d["inherited_evidence"]] == [eid]
    assert d["evidence"] == [], "inherited evidence must NOT leak into the direct list"

    # a question mapped to a different control does not inherit it
    d2 = app_client.get(f"/api/assessments/{aid}/responses/{ids['q2']}", headers=h).json()
    assert d2["inherited_evidence"] == []


def test_remapping_changes_which_evidence_is_inherited(app_client):
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-inh2")
    aid = _assessment(app_client, h, ids["tpl"])
    ev = app_client.post("/api/evidence", headers=h,
                         data={"title": "MFA screenshot", "evidence_type": "SCREENSHOT"},
                         files={"file": ("mfa.png", b"\x89PNG x", "image/png")}).json()["id"]
    app_client.post(f"/api/library/controls/{ids['ctl2']}/evidence", headers=h,
                    json={"evidence_id": ev})

    before = app_client.get(f"/api/assessments/{aid}/responses/{ids['q1']}", headers=h).json()
    assert before["inherited_evidence"] == []

    app_client.patch(f"/api/assessments/{aid}/responses/{ids['q1']}/mapping",
                     headers=h, json={"control_id": ids["ctl2"]})
    after = app_client.get(f"/api/assessments/{aid}/responses/{ids['q1']}", headers=h).json()
    assert [e["id"] for e in after["inherited_evidence"]] == [ev]


# ---------------------------------------------------------------- grid + export

def test_grid_carries_the_mapped_control_statement(app_client):
    """The grid showed only the control CODE; the statement is what tells a reviewer
    whether the mapping is actually right."""
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-grid")
    aid = _assessment(app_client, h, ids["tpl"])
    grid = {r["number"]: r for r in
            app_client.get(f"/api/assessments/{aid}/questions", headers=h).json()}
    assert grid["1"]["mapped_control"] == ids["code1"]
    assert grid["1"]["mapped_control_statement"] == "ISP"


def test_export_names_the_evidence_file_not_just_its_title(app_client):
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-exp")
    aid = _assessment(app_client, h, ids["tpl"])
    app_client.put(f"/api/assessments/{aid}/responses/{ids['q1']}", headers=h,
                   json={"response_value": "yes", "comment": "done"})
    ev = app_client.post("/api/evidence", headers=h,
                         data={"title": "Access review", "evidence_type": "REPORT"},
                         files={"file": ("access-review-q3.xlsx", b"PK\x03\x04x",
                                         "application/vnd.ms-excel")}).json()["id"]
    app_client.post(f"/api/assessments/{aid}/responses/{ids['q1']}/evidence", headers=h,
                    json={"evidence_id": ev})

    r = app_client.get(f"/api/assessments/{aid}/export.xlsx", headers=h)
    assert r.status_code == 200 and r.content[:2] == b"PK"

    import io
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(r.content)).active
    cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert any("Access review" in v and "access-review-q3.xlsx" in v for v in cells), \
        "the export must name the actual file, not only the evidence title"


def test_export_is_scoped_to_this_assessment(app_client):
    """The evidence query used to have no WHERE at all — it walked every
    response_evidence row in the database. A second assessment's evidence must not
    appear in this one's export."""
    from api.database import engine
    h = _h(app_client)
    a_ids = _seed(engine, _tid(engine), suffix="-sc1")
    b_ids = _seed(engine, _tid(engine), suffix="-sc2")
    a_aid, b_aid = _assessment(app_client, h, a_ids["tpl"]), _assessment(app_client, h, b_ids["tpl"])

    app_client.put(f"/api/assessments/{b_aid}/responses/{b_ids['q1']}", headers=h,
                   json={"response_value": "yes"})
    other = app_client.post("/api/evidence", headers=h,
                            data={"title": "OTHER-ASSESSMENT-ONLY", "evidence_type": "REPORT"},
                            files={"file": ("other.txt", b"x", "text/plain")}).json()["id"]
    app_client.post(f"/api/assessments/{b_aid}/responses/{b_ids['q1']}/evidence", headers=h,
                    json={"evidence_id": other})

    app_client.put(f"/api/assessments/{a_aid}/responses/{a_ids['q1']}", headers=h,
                   json={"response_value": "yes"})
    r = app_client.get(f"/api/assessments/{a_aid}/export.xlsx", headers=h)
    import io
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(r.content)).active
    cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert not any("OTHER-ASSESSMENT-ONLY" in v for v in cells)


def test_export_survives_an_assessment_with_no_responses(app_client):
    """The `if resp_ids:` guard — an empty IN () is a SQL error in some dialects and an
    always-false predicate in others; either way the export must not crash."""
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-empty")
    aid = _assessment(app_client, h, ids["tpl"])
    r = app_client.get(f"/api/assessments/{aid}/export.xlsx", headers=h)
    assert r.status_code == 200 and r.content[:2] == b"PK"
    grid = app_client.get(f"/api/assessments/{aid}/questions", headers=h)
    assert grid.status_code == 200


def test_remap_supersedes_the_previous_mapping(app_client):
    """The subtle one. A remap ADDS a row to question_control_map; without explicitly
    rejecting the old one it survives as a competing mapping, and _best_controls — which
    ranks confirmed rows by confidence — keeps preferring the importer's original (which
    has a real confidence score) over the human's hand-picked one (which has none). The
    remap would appear to succeed and change nothing."""
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-sup")
    aid = _assessment(app_client, h, ids["tpl"])

    app_client.patch(f"/api/assessments/{aid}/responses/{ids['q1']}/mapping",
                     headers=h, json={"control_id": ids["ctl2"]})

    with engine.connect() as c:
        rows = dict(c.execute(sqltext(
            "SELECT control_id, status FROM question_control_map WHERE question_id=:q"),
            {"q": ids["q1"]}).all())
    assert rows[ids["ctl1"]] == "rejected", "the old mapping must be superseded, not left live"
    assert rows[ids["ctl2"]] == "confirmed"

    # and the grid agrees — this is what the reviewer actually sees
    grid = {r["number"]: r for r in
            app_client.get(f"/api/assessments/{aid}/questions", headers=h).json()}
    assert grid["1"]["mapped_control"] == ids["code2"]


def test_remapping_back_to_the_original_control_works(app_client):
    """Remap A->B->A. The A row is 'rejected' by the first remap, so the second must
    revive it rather than leaving the question with no live mapping at all."""
    from api.database import engine
    h = _h(app_client)
    ids = _seed(engine, _tid(engine), suffix="-back")
    aid = _assessment(app_client, h, ids["tpl"])

    app_client.patch(f"/api/assessments/{aid}/responses/{ids['q1']}/mapping",
                     headers=h, json={"control_id": ids["ctl2"]})
    r = app_client.patch(f"/api/assessments/{aid}/responses/{ids['q1']}/mapping",
                         headers=h, json={"control_id": ids["ctl1"]})
    assert r.status_code == 200

    d = app_client.get(f"/api/assessments/{aid}/responses/{ids['q1']}", headers=h).json()
    assert d["mapped_control"]["code"] == ids["code1"]
