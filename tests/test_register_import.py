import io
import json
import uuid

from openpyxl import load_workbook

from tests.test_registers import _h


def _person(app_client, h, name, email=None):
    return app_client.post("/api/people", headers=h, json={
        "full_name": name, "email": email or f"{uuid.uuid4().hex[:8]}@kiam.example"}).json()["id"]


def _upload(app_client, h, register, content, name="import.csv", mapping=None):
    data = {"mapping": json.dumps(mapping)} if mapping else None
    return app_client.post(f"/api/import/{register}", headers=h, data=data,
                           files={"file": (name, content, "text/csv")})


def test_template_downloads_and_round_trips(app_client):
    h = _h(app_client)
    r = app_client.get("/api/import/risks/template.xlsx", headers=h)
    assert r.status_code == 200 and r.content[:2] == b"PK"
    wb = load_workbook(io.BytesIO(r.content))
    assert "How to fill this in" in wb.sheetnames
    # an UNFILLED template must import as zero rows, not as its own guidance
    up = _upload(app_client, h, "risks", r.content, "t.xlsx")
    assert up.status_code == 400 and "no data rows" in up.json()["detail"]


def test_a_good_file_imports_every_row(app_client):
    h = _h(app_client)
    tag = uuid.uuid4().hex[:6]
    # lower-case on the second row on purpose: spreadsheets carry any casing, and
    # `one_of` normalises to the canonical spelling the CHECK constraint wants.
    csv = (f"Title,Reference,Treatment\n"
           f"{tag} one,R-1,MITIGATED\n{tag} two,R-2,accepted\n").encode()
    r = _upload(app_client, h, "risks", csv)
    assert r.status_code == 200, r.text
    assert r.json() == {"created": 2, "failed": 0, "errors": []}
    titles = [x["title"] for x in app_client.get("/api/risks", headers=h).json()]
    assert f"{tag} one" in titles and f"{tag} two" in titles


def test_one_bad_row_does_not_kill_the_batch(app_client):
    h = _h(app_client)
    tag = uuid.uuid4().hex[:6]
    csv = (f"Title,Inherent likelihood\n"
           f"{tag} good,3\n"
           f",4\n"                       # missing required title
           f"{tag} bad score,99\n"       # out of range
           f"{tag} also good,1\n").encode()
    r = _upload(app_client, h, "risks", csv)
    body = r.json()
    assert body["created"] == 2, body
    assert body["failed"] == 2
    rows = {e["row"] for e in body["errors"]}
    assert rows == {3, 4}, body["errors"]          # excel row numbers, header is row 1
    assert any("required" in e["error"] for e in body["errors"])
    assert any("1 to 5" in e["error"] for e in body["errors"])


def test_owner_resolves_by_name_and_by_email(app_client):
    h = _h(app_client)
    tag = uuid.uuid4().hex[:6]
    email = f"solo-{tag}@kiam.example"
    _person(app_client, h, f"Solo Owner {tag}", email)
    csv = (f"Title,Owner\n{tag} by-name,Solo Owner {tag}\n{tag} by-email,{email}\n").encode()
    r = _upload(app_client, h, "risks", csv)
    assert r.json()["created"] == 2, r.json()
    owners = {x["title"]: x["owner_name"] for x in app_client.get("/api/risks", headers=h).json()}
    assert owners[f"{tag} by-name"] == f"Solo Owner {tag}"
    assert owners[f"{tag} by-email"] == f"Solo Owner {tag}"


def test_an_ambiguous_owner_fails_the_row_rather_than_guessing(app_client):
    """The rule that matters most: picking arbitrarily between two people with the same name
    assigns the risk to the wrong owner AND reports success."""
    h = _h(app_client)
    tag = uuid.uuid4().hex[:6]
    dupe = f"Twin Person {tag}"
    _person(app_client, h, dupe)
    _person(app_client, h, dupe)
    csv = f"Title,Owner\n{tag} ambiguous,{dupe}\n".encode()
    body = _upload(app_client, h, "risks", csv).json()
    assert body["created"] == 0 and body["failed"] == 1
    assert "more than one" in body["errors"][0]["error"]
    assert "email" in body["errors"][0]["error"]


def test_an_unknown_owner_fails_the_row_naming_the_value(app_client):
    h = _h(app_client)
    body = _upload(app_client, h, "risks", b"Title,Owner\nX,Nobody At All\n").json()
    assert body["failed"] == 1
    assert "Nobody At All" in body["errors"][0]["error"]


def test_explicit_mapping_beats_the_default_labels(app_client):
    h = _h(app_client)
    tag = uuid.uuid4().hex[:6]
    csv = f"Risk name,Our ref\n{tag} mapped,X-9\n".encode()
    # no header matches our labels, so without a mapping the required column is unmapped
    assert _upload(app_client, h, "risks", csv).status_code == 400
    r = _upload(app_client, h, "risks", csv,
                mapping={"title": "Risk name", "reference": "Our ref"})
    assert r.status_code == 200 and r.json()["created"] == 1, r.text


def test_every_register_has_a_template_and_columns(app_client):
    h = _h(app_client)
    for reg in ("risks", "assets", "data-items", "third-parties", "incidents"):
        cols = app_client.get(f"/api/import/{reg}/columns", headers=h)
        assert cols.status_code == 200, reg
        assert any(c["required"] for c in cols.json()["columns"]), reg
        tpl = app_client.get(f"/api/import/{reg}/template.xlsx", headers=h)
        assert tpl.status_code == 200 and tpl.content[:2] == b"PK", reg


def test_unknown_register_is_404(app_client):
    h = _h(app_client)
    assert app_client.get("/api/import/nonsense/columns", headers=h).status_code == 404


def test_vendor_reference_resolves_for_assets(app_client):
    h = _h(app_client)
    tag = uuid.uuid4().hex[:6]
    vendor = f"Acme {tag}"
    app_client.post("/api/third-parties", headers=h, json={"name": vendor})
    csv = f"Name,Vendor\n{tag} laptop,{vendor}\n".encode()
    r = _upload(app_client, h, "assets", csv)
    assert r.json()["created"] == 1, r.json()


def test_a_bad_enum_is_a_row_error_not_a_database_error(app_client):
    h = _h(app_client)
    body = _upload(app_client, h, "assets", b"Name,Criticality\nThing,VERY HIGH\n").json()
    assert body["failed"] == 1
    assert "LOW, MEDIUM, HIGH, CRITICAL" in body["errors"][0]["error"]


def test_people_import_now_accepts_xlsx_too(app_client):
    """P5-S5 moved people onto the shared reader; it was CSV-only and died on a .xlsx with an
    opaque BadZipFile-shaped error."""
    from openpyxl import Workbook
    h = _h(app_client)
    tag = uuid.uuid4().hex[:6]
    wb = Workbook(); ws = wb.active
    ws.append(["full_name", "email"])
    ws.append([f"Xlsx Person {tag}", f"xp-{tag}@kiam.example"])
    buf = io.BytesIO(); wb.save(buf)

    r = app_client.post("/api/people/import", headers=h,
                        files={"file": ("people.xlsx", buf.getvalue(),
                                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    assert r.json()["created"] == 1, r.json()
    names = [p["full_name"] for p in app_client.get("/api/people", headers=h).json()]
    assert f"Xlsx Person {tag}" in names


def test_people_import_still_takes_csv_and_reports_bad_rows(app_client):
    h = _h(app_client)
    tag = uuid.uuid4().hex[:6]
    csv = (f"full_name,email\n"
           f"Good {tag},good-{tag}@kiam.example\n"
           f",no-name-{tag}@kiam.example\n").encode()
    body = app_client.post("/api/people/import", headers=h,
                           files={"file": ("p.csv", csv, "text/csv")}).json()
    assert body["created"] == 1 and body["failed"] == 1, body
    assert body["errors"][0]["row"] == 3
