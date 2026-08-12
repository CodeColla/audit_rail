"""XLSX import parsing and export building (M6), via openpyxl.

Import: generalizes the Phase-0 stdlib parser — detect a header row and the
question / number / section columns (or take them as hints), then yield rows.
Export: build a clean answers workbook from an assessment's responses.
"""

from __future__ import annotations

import csv
import io
import re

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

_QUESTION_HINTS = ("question", "control", "requirement", "particular", "checkpoint")
_NUMBER_HINTS = ("s.no", "sr.no", "sno", "srno", "s no", "no.", "ref", "#", "control no")
_SECTION_HINTS = ("domain", "section", "category", "area", "control area", "sub domain")


def _norm(v) -> str:
    return re.sub(r"\s+", " ", str(v).strip().lower()) if v is not None else ""


def _header_role(val: str):
    """(role, strength) for a header cell. 'question' with strength 2 = a column
    literally named 'question' (beats a weak 'control'/'requirement' match, and
    keeps 'Control Domain' as a section rather than the question column)."""
    if "question" in val:
        return "question", 2
    if any(h in val for h in _SECTION_HINTS):
        return "section", 0
    if any(val == h or h in val for h in _NUMBER_HINTS):
        return "number", 0
    if any(h in val for h in _QUESTION_HINTS):  # control/requirement/... = weak question
        return "question", 1
    return None, 0


_INTERROGATIVE = (
    "do ", "does ", "did ", "is ", "are ", "was ", "were ", "has ", "have ", "how ",
    "what ", "which ", "when ", "where ", "why ", "who ", "can ", "could ", "would ",
    "list ", "describe ", "provide ", "explain ", "specify ", "state ", "confirm ",
    "share ", "detail ", "mention ",
)


def _question_score(v) -> float:
    """How much a cell reads like a CHECKLIST QUESTION rather than an answer or a label.

    Length alone is not enough: a vendor's answer ("Security policies are reviewed
    annually…") is just as long as the question it answers, and picking the answer column
    imports the wrong text entirely. Interrogative shape is what separates them.
    """
    s = str(v).strip() if v is not None else ""
    if len(s) < 12 or " " not in s:
        return 0.0
    if s.endswith("?"):
        return 1.0
    if s.lower().startswith(_INTERROGATIVE):
        return 0.9
    return 0.35 if len(s) >= 25 else 0.0        # prose, but not obviously a question


def _column_quality(ws, header_row: int, col: int, sample: int = 30) -> float:
    """Fraction of the cells BELOW a header that actually read like questions.

    Header words alone are not enough to identify the question column: banks label answer
    columns things like "Control Applicability (Yes/No)", which matches the 'control' hint
    but is full of "Yes". Looking at the data underneath is what tells them apart.
    """
    seen, score = 0, 0.0
    for r in range(header_row + 1, min(header_row + 1 + sample, (ws.max_row or 0) + 1)):
        row = ws[r]
        v = row[col].value if col < len(row) else None
        if v is None or not str(v).strip():
            continue
        seen += 1
        score += _question_score(v)
    return score / seen if seen else 0.0


def _find_header(ws, scan_rows: int = 12, min_quality: float = 0.5):
    """Return (header_row_idx, {role: col_idx}) by scanning the first rows.

    Two passes: first trust the header words, but only if the column beneath them really
    holds questions; otherwise fall back to whichever column in the scan window has the
    most question-like content. Without the second pass, sheets whose real header sits
    below a title row (or which have no header at all) imported 'S.No.' / 'Yes' as the
    question text.
    """
    best_fallback = (0.0, None, None)          # (quality, row, col)

    for r in range(1, min(scan_rows, ws.max_row or 1) + 1):
        cells = [_norm(c.value) for c in ws[r]]
        # track the best-looking data column under this row, for the fallback pass
        for i in range(len(cells)):
            qual = _column_quality(ws, r, i)
            if qual > best_fallback[0]:
                best_fallback = (qual, r, i)

        joined = " | ".join(cells)
        if not any(h in joined for h in _QUESTION_HINTS):
            continue
        q_best, q_strength = None, -1
        cols: dict = {}
        for i, val in enumerate(cells):
            if not val:
                continue
            role, strength = _header_role(val)
            if role == "question":
                # prefer a stronger header word, but only among columns that hold questions
                if strength > q_strength and _column_quality(ws, r, i) >= min_quality:
                    q_best, q_strength = i, strength
            elif role and role not in cols:
                cols[role] = i
        if q_best is not None:
            cols["question"] = q_best
            return r, cols

    qual, r, i = best_fallback
    if r is not None and qual >= min_quality:
        return r, {"question": i}
    return None, {}


def _is_csv(data: bytes) -> bool:
    """XLSX is a zip archive, so it always starts with 'PK'. Anything else we treat as text."""
    return not data[:2] == b"PK"


def _workbook_from_csv(data: bytes):
    """Turn CSV bytes into an in-memory workbook so ONE parsing path serves both formats.

    The UI has always advertised `accept=".xlsx,.csv"`, but the parser was openpyxl-only
    and a CSV died with an opaque BadZipFile. Sniffing the delimiter handles the
    comma/semicolon/tab variants banks actually send.
    """
    text = data.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    wb = Workbook()
    ws = wb.active
    for row in csv.reader(io.StringIO(text), dialect):
        ws.append(row if row else [None])
    return wb


def read_rows(data: bytes, sheet: str | None = None) -> tuple[list[str], list[dict]]:
    """Any .xlsx or .csv -> (headers, rows-as-dicts). The generic reader behind bulk import.

    Deliberately dumb: it takes the FIRST row as the header and makes no attempt to guess
    which column means what. That guessing is `parse_checklist`'s job and it exists because a
    bank's checklist arrives in whatever shape the bank chose. A register import is the other
    situation — the user is filling in OUR template, or telling us the mapping explicitly in
    the UI — so inferring here would only add a way to be confidently wrong.

    Reuses `_is_csv`/`_workbook_from_csv`, so one path serves both formats exactly as the
    checklist importer already does.
    """
    if _is_csv(data):
        wb = _workbook_from_csv(data)
        ws = wb.active
    else:
        wb = load_workbook(io.BytesIO(data), read_only=False, data_only=True)
        if sheet:
            if sheet not in wb.sheetnames:
                raise ValueError(
                    f"no sheet named {sheet!r}; this file has: {', '.join(wb.sheetnames)}")
            ws = wb[sheet]
        else:
            ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    if not any(headers):
        raise ValueError("the first row must be a header row naming the columns")

    out: list[dict] = []
    for values in rows_iter:
        row = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = values[i] if i < len(values) else None
            row[h] = str(v).strip() if v is not None and str(v).strip() != "" else None
        # A wholly blank line is padding, not a row the user meant to import — reporting it
        # as an error would bury the real failures under trailing-whitespace noise.
        if any(v is not None for v in row.values()):
            out.append(row)
    return headers, out


def build_template(columns: list[dict], sheet_name: str = "Import") -> bytes:
    """A one-sheet .xlsx whose header row is exactly what the importer expects.

    `columns` = [{"key", "label", "help", "required"}].

    Guidance lives on a SECOND sheet, not in a second header row. An earlier version put the
    hints directly under the headers, which reads nicely and then imports as a junk record
    the moment the user fills the file in and sends it back — `read_rows` cannot tell a hint
    from data. Headers are the bare labels for the same reason: a " *" suffix marking
    "required" would have to be stripped before mapping, and something eventually forgets to.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Import"
    ws.append([c["label"] for c in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for i, c in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(40, len(c["label"]) + 8))

    guide = wb.create_sheet("How to fill this in")
    guide.append(["Column", "Required", "What to put in it"])
    for cell in guide[1]:
        cell.font = Font(bold=True)
    for c in columns:
        guide.append([c["label"], "yes" if c.get("required") else "optional", c.get("help", "")])
    for col, width in (("A", 26), ("B", 12), ("C", 64)):
        guide.column_dimensions[col].width = width
    guide["C1"].alignment = Alignment(horizontal="left")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_checklist(data: bytes, sheet: str | None = None,
                    question_col: int | None = None, number_col: int | None = None,
                    section_col: int | None = None, header_row: int | None = None):
    """Return (meta, rows). rows = [{number, section, text}] with section forward-filled."""
    if _is_csv(data):
        wb = _workbook_from_csv(data)
    else:
        # not read_only: in read_only mode ws.max_row can be None for some files
        wb = load_workbook(io.BytesIO(data), read_only=False, data_only=True)

    if sheet:
        if sheet not in wb.sheetnames:
            # Silently falling back to the active sheet imported the WRONG data on a typo.
            raise ValueError(
                f"no sheet named {sheet!r}; this file has: {', '.join(wb.sheetnames)}")
        ws = wb[sheet]
    else:
        ws = wb.active

    if question_col is None:
        header_row, cols = _find_header(ws)
        if header_row is None and not sheet:
            # The active sheet is often a leftover empty "Sheet2" — the real checklist is
            # on another tab. Try them all before giving up.
            for cand in wb.worksheets:
                if cand is ws:
                    continue
                hr, c = _find_header(cand)
                if hr is not None:
                    ws, header_row, cols = cand, hr, c
                    break
        if header_row is None:
            raise ValueError(
                "could not detect a question column — pass question_col (a column letter), "
                f"or name the sheet. Sheets in this file: {', '.join(wb.sheetnames)}")
    else:
        cols = {"question": question_col}
        if number_col is not None:
            cols["number"] = number_col
        if section_col is not None:
            cols["section"] = section_col
        header_row = header_row or 1

    rows, section = [], ""
    for r in range(header_row + 1, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        def get(role):
            i = cols.get(role)
            return row[i] if i is not None and i < len(row) else None
        if cols.get("section") is not None and get("section"):
            section = str(get("section")).strip()
        q = get("question")
        if q is None or not str(q).strip():
            continue
        num = get("number")
        rows.append({
            "number": ("" if num is None else str(num).strip()),
            "section": section,
            "text": str(q).strip(),
        })
    meta = {"sheet": ws.title, "header_row": header_row, "columns": cols,
            "row_count": len(rows)}
    return meta, rows


def build_sheet_workbook(title: str, content: str) -> bytes:
    """A SHEET document version -> a real .xlsx (P5-S2b).

    Jspreadsheet CE cannot write .xlsx — that is a paid-tier feature — so we build it here
    from our own stored format with openpyxl, which the backend already depends on. That
    also keeps the file correct regardless of what the browser grid does next; see
    docs/phase5/04-spreadsheet-library-evaluation.md.

    **Formulas are written as formulas, not as their frozen values.** An .xlsx is a working
    copy — someone who opens it expects to change an input and see totals move. The
    controlled, signed artefact remains the PDF, whose numbers are frozen at publish. The
    consequence, stated plainly because it is a real trade-off: a recipient who edits the
    spreadsheet will see figures that no longer match the signed PDF, which is exactly what
    should happen to a working copy.
    """
    from api import render  # noqa: PLC0415 — avoids a circular import at module load

    book = render.parse_sheet(content or "")
    wb = Workbook()
    wb.remove(wb.active)                     # drop openpyxl's default empty sheet

    for sheet in book["sheets"]:
        # Excel forbids []:*?/\ in sheet names and caps them at 31 chars. A rejected name
        # raises deep inside openpyxl on save, so sanitise rather than propagate.
        safe = re.sub(r"[\[\]:*?/\\]", "-", sheet["name"])[:31] or "Sheet"
        ws = wb.create_sheet(title=safe)
        for r, row in enumerate(sheet["data"], start=1):
            for c, value in enumerate(row, start=1):
                formula = sheet["formulas"].get(f"{render._col_letter(c - 1)}{r}")
                cell = ws.cell(row=r, column=c)
                # Write numbers as numbers — a register exported with every figure stored as
                # text is useless for sorting or further calculation in Excel.
                cell.value = formula if formula else _coerce(value)
                props = sheet["style"].get(f"{render._col_letter(c - 1)}{r}")
                wrapped = bool(sheet.get("colWrap") and c - 1 < len(sheet["colWrap"])
                               and sheet["colWrap"][c - 1])
                if props or wrapped:
                    _apply_style(cell, props or {}, wrap=wrapped)
                # A formatted column travels as a NUMBER carrying an Excel number format,
                # never as pre-formatted text: "₹1,234.50" as a string is unsortable and
                # uncomputable, which defeats the point of exporting a working copy. Applied
                # only where the cell really is numeric or a formula, mirroring
                # `render.format_cell_value` leaving text alone — so a "Amount" header in a
                # currency column stays plain.
                fmts = sheet.get("colFormat") or []
                number_format = _number_format(fmts[c - 1] if c - 1 < len(fmts) else "")
                if number_format and (formula or isinstance(cell.value, (int, float))):
                    cell.number_format = number_format
                # A reviewer's note becomes a real Excel comment (P6-S5b), not a suffix on the
                # value — so it stays a note rather than becoming part of the data, and Excel
                # shows it the same way it shows its own.
                note = (sheet.get("comments") or {}).get(
                    f"{render._col_letter(c - 1)}{r}")
                if note:
                    cell.comment = Comment(note, "Audit Rail")

        for addr, span in sheet["merges"].items():
            m = re.match(r"^([A-Z]+)(\d+)$", addr)
            if not m:
                continue
            c0 = column_index_from_string(m.group(1))
            r0 = int(m.group(2))
            if span["colspan"] > 1 or span["rowspan"] > 1:
                ws.merge_cells(start_row=r0, start_column=c0,
                               end_row=r0 + span["rowspan"] - 1,
                               end_column=c0 + span["colspan"] - 1)

        for i, width in enumerate(sheet["colWidths"], start=1):
            # jspreadsheet stores pixels; Excel's unit is roughly one character (~7px).
            ws.column_dimensions[get_column_letter(i)].width = max(2.0, width / 7.0)

    if not wb.sheetnames:                    # a document with no worksheets at all
        wb.create_sheet(title="Sheet1")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _number_format(fmt: str) -> str | None:
    """Our column format -> an Excel number-format string (P6-S4).

    `0.00%` multiplies by 100 on display, which is why `render.format_cell_value` treats a
    stored percent as a ratio: any other choice would make the .xlsx and the PDF show
    different numbers for the same cell. The currency symbol is quoted so Excel treats it as
    a literal rather than trying to parse it as a format token."""
    from api import render  # noqa: PLC0415 — same circular-import dodge as the caller

    if fmt == "percent":
        return "0.00%"
    if fmt.startswith("currency:"):
        symbol = render.CURRENCY_SYMBOLS.get(fmt.split(":", 1)[1])
        if symbol:
            return f'"{symbol}"#,##0.00'
    return None


def _coerce(value: str):
    """Store a numeric-looking cell as a number so Excel can compute with it."""
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        pass
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _apply_style(cell, props: dict, wrap: bool = False) -> None:
    """Map our validated style properties onto openpyxl. Values were already checked by
    `render._STYLE_PROPS`, so hex colours are known-good; openpyxl wants them without '#'
    and prefixed with an alpha byte."""
    if any(props.get(k) for k in ("bold", "italic", "underline", "fontSize", "color")):
        cell.font = Font(
            bold=bool(props.get("bold")),
            italic=bool(props.get("italic")),
            underline="single" if props.get("underline") else None,
            size=props.get("fontSize") or None,
            color=f"FF{props['color'].lstrip('#').upper()}" if props.get("color") else None)
    if props.get("align") or wrap:
        # Excel needs wrap_text on the Alignment object, so it travels with horizontal
        # alignment rather than being a separate attribute.
        cell.alignment = Alignment(horizontal=props.get("align") or None,
                                   wrap_text=True if wrap else None)
    if props.get("background"):
        fill = f"FF{props['background'].lstrip('#').upper()}"
        cell.fill = PatternFill(fill_type="solid", start_color=fill, end_color=fill)


def build_answers_workbook(assessment: dict, rows: list[dict]) -> bytes:
    """rows = [{number, section, text, response_value, comment, final_status, evidence}]."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Responses"
    ws.append([f"Assessment: {assessment.get('title', '')}",
               f"Bank: {assessment.get('bank_name', '')}",
               f"Status: {assessment.get('status', '')}"])
    ws.append([])
    headers = ["S.No", "Section", "Control Question", "Response", "Comments",
               "Status", "Evidence"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get("number", ""), row.get("section", ""), row.get("text", ""),
                   (row.get("response_value") or "").upper(), row.get("comment") or "",
                   row.get("final_status") or row.get("workflow_status") or "",
                   row.get("evidence") or ""])
    widths = [10, 22, 70, 12, 40, 16, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
